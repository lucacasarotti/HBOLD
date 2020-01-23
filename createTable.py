from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from extractor.util import mongo
from datetime import date
import pymongo as pm
import os.path


client = pm.MongoClient()
dbLodex=client.lodex



def countDataset():
    count = 0
    for end in mongo.getAllEndopoinLodex():
        count = count + 1
    return count


def countExt_month_year(month, year):
    count = 0
    for end in mongo.getAllEndopoinLodex():
        p = mongo.getExtById(end['_id'])
        if len(p) != 0:
            e = mongo.getLastRunById(end['_id'])
            if (e['date'].month == month) and (e['date'].year == year):
                count = count + 1
    return count


def countExt():
    count = 0
    for end in mongo.getAllEndopoinLodex():
        p = mongo.getExtById(end['_id'])
        if len(p) != 0:
            count = count + 1
    return count


def countIke_month_year(month, year):
    count=0
    for e in dbLodex.ike.find({}):
        if (e['ikeDate'].month == month) and (e['ikeDate'].year == year):
            count = count + 1
    return count


def countIke():
    count=0
    for e in dbLodex.ike.find({}):
        count = count + 1
    return count


def main():
    workbook_name = 'table.xlsx'
    page = 0

    if not os.path.exists(workbook_name):
        headers = ['Date','Dataset URLs','Dataset with successfully EXT in this month','Dataset with successfully EXT','Dataset with successfully IKE in this month', 'Dataset with successfully IKE']
        wb = Workbook()
        page = wb.active
        page.title = 'info'
        page.append(headers) #write the headers to the first line
        column = 1
        while column <= 6:
            i = get_column_letter(column)
            page.column_dimensions[i].width = 35.91
            column += 1
    else:
        wb = load_workbook(workbook_name)
        page = wb.active

    i0 = date.today()
    i1 = countDataset()
    i2 = countExt_month_year(i0.month, i0.year)
    i3 = countExt()
    i4 = countIke_month_year(i0.month, i0.year)
    i5 = countIke()

    info = [i0, i1, i2, i3, i4, i5]
    page.append(info)

    wb.save(filename=workbook_name)


if __name__ == "__main__":
    main()

